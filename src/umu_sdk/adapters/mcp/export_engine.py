# umu-skills: unofficial UMU platform automation helpers
# This file is part of an independent, third-party project and is not
# affiliated with UMU. Use at your own risk.

"""通用导出引擎.

为 UMU MCP Server 提供查询结果导出能力，支持 Excel/CSV 格式。
当前实现包含课程访问权限导出模板，未来可扩展更多场景。
"""

from __future__ import annotations

import csv
import json
import os
import time
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from ...core.client import UMUClient
from .shared_access_permissions import (
    _get_obj_access_list,
    _get_obj_access_permission,
    _permission_text,
)


class ExportEngine:
    """通用导出引擎，用于将 UMU 查询结果导出为 Excel/CSV."""

    def __init__(self, client: UMUClient) -> None:
        """初始化导出引擎.

        Args:
            client: 已登录的 UMUClient 实例。
        """
        self.client = client

    # -----------------------------------------------------------------------
    # 通用写入接口
    # -----------------------------------------------------------------------
    def to_excel(
        self,
        records: list[dict[str, Any]],
        output_path: str,
        sheet_name: str = "Sheet1",
    ) -> str:
        """将记录列表导出为 Excel 文件.

        Args:
            records: 字典列表，每个字典表示一行。
            output_path: 输出文件路径。
            sheet_name: 工作表名称。

        Returns:
            输出文件路径。
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws.title = sheet_name

        if not records:
            wb.save(output_path)
            return output_path

        headers = list(records[0].keys())
        ws.append(headers)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for record in records:
            ws.append([record.get(h, "") for h in headers])

        # 自动调整列宽，上限 60
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        wb.save(output_path)
        return output_path

    def to_csv(
        self,
        records: list[dict[str, Any]],
        output_path: str,
    ) -> str:
        """将记录列表导出为 CSV 文件.

        Args:
            records: 字典列表。
            output_path: 输出文件路径。

        Returns:
            输出文件路径。
        """
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            if not records:
                return output_path
            writer = csv.DictWriter(f, fieldnames=list(records[0].keys()))
            writer.writeheader()
            writer.writerows(records)
        return output_path

    def export_records(
        self,
        records: list[dict[str, Any]],
        output_path: str,
        sheet_name: str = "Sheet1",
    ) -> dict[str, Any]:
        """通用兜底接口：将任意记录列表导出为 Excel/CSV.

        Args:
            records: 字典列表，每个字典表示一行。
            output_path: 输出文件路径，根据扩展名自动选择 xlsx/csv。
            sheet_name: Excel 工作表名称（CSV 时忽略）。

        Returns:
            包含 file_path、total_records 的字典。
        """
        ext = os.path.splitext(output_path)[1].lower()
        if ext == ".csv":
            self.to_csv(records, output_path)
        else:
            self.to_excel(records, output_path, sheet_name=sheet_name)

        return {
            "file_path": output_path,
            "total_records": len(records),
        }

    # -----------------------------------------------------------------------
    # 场景模板：课程访问权限导出
    # -----------------------------------------------------------------------
    def export_course_permissions(
        self,
        output_path: str,
        *,
        include_access_list: bool = True,
    ) -> dict[str, Any]:
        """导出当前讲师创建的所有课程的访问权限明细.

        Args:
            output_path: 输出文件路径，根据扩展名自动选择 xlsx/csv。
            include_access_list: 是否展开指定账户的授权列表。

        Returns:
            包含 file_path、total_courses、total_records 的字典。
        """
        courses = self._fetch_all_created_courses()
        records: list[dict[str, Any]] = []

        for course in courses:
            group_id = course.get("group_id", "")
            title = course.get("title", "")
            access_code = course.get("access_code", "")

            try:
                access_permission, _, _ = _get_obj_access_permission(
                    self.client, group_id, "group"
                )
                permission_text = _permission_text(access_permission)
            except Exception as e:
                records.append(
                    {
                        "group_id": group_id,
                        "title": title,
                        "access_code": access_code,
                        "access_permission": "",
                        "permission_text": f"获取失败: {e}",
                        "authorized_account": "",
                        "authorized_account_type": "",
                        "authorized_account_id": "",
                        "authorized_account_name": "",
                    }
                )
                continue

            if access_permission == 3 and include_access_list:
                accounts = self._fetch_all_access_list(group_id, "group")
                if accounts:
                    for account in accounts:
                        records.append(
                            {
                                "group_id": group_id,
                                "title": title,
                                "access_code": access_code,
                                "access_permission": access_permission,
                                "permission_text": permission_text,
                                "authorized_account": account.get("account", ""),
                                "authorized_account_type": account.get("account_type", ""),
                                "authorized_account_id": account.get("id", ""),
                                "authorized_account_name": self._account_display_name(account),
                            }
                        )
                else:
                    records.append(
                        {
                            "group_id": group_id,
                            "title": title,
                            "access_code": access_code,
                            "access_permission": access_permission,
                            "permission_text": permission_text,
                            "authorized_account": "",
                            "authorized_account_type": "",
                            "authorized_account_id": "",
                            "authorized_account_name": "",
                        }
                    )
            else:
                records.append(
                    {
                        "group_id": group_id,
                        "title": title,
                        "access_code": access_code,
                        "access_permission": access_permission,
                        "permission_text": permission_text,
                        "authorized_account": "",
                        "authorized_account_type": "",
                        "authorized_account_id": "",
                        "authorized_account_name": "",
                    }
                )

        ext = os.path.splitext(output_path)[1].lower()
        if ext == ".csv":
            self.to_csv(records, output_path)
        else:
            self.to_excel(records, output_path, sheet_name="课程访问权限")

        return {
            "file_path": output_path,
            "total_courses": len(courses),
            "total_records": len(records),
        }

    # -----------------------------------------------------------------------
    # 内部辅助方法
    # -----------------------------------------------------------------------
    def _fetch_all_created_courses(self) -> list[dict[str, Any]]:
        """分页获取当前讲师创建的所有课程."""
        all_items: list[dict[str, Any]] = []
        page = 1
        size = 50
        total = 0

        while True:
            resp = self.client.get(
                self.client.desktop_url("/api/group/getgrouplist"),
                params={
                    "t": str(int(time.time() * 1000)),
                    "from_type": "web",
                    "order": "update_time",
                    "page": str(page),
                    "size": str(size),
                },
            )
            if resp.get("status") not in (True, "true") and resp.get("error_code") != 0:
                raise RuntimeError(resp.get("error", "获取已创建课程列表失败"))

            data = resp.get("data", {})
            page_info = data.get("page_info", {})
            course_list = data.get("list", [])
            total = int(page_info.get("list_total_num", 0) or total)

            for item in course_list:
                info = item.get("groupInfo", {})
                all_items.append(
                    {
                        "group_id": info.get("id", ""),
                        "title": info.get("title", ""),
                        "access_code": info.get("access_code", ""),
                    }
                )

            if not course_list or len(all_items) >= total:
                break
            page += 1
            if page > 50:
                break

        return all_items

    def _fetch_all_access_list(
        self,
        obj_id: str,
        obj_type: str,
    ) -> list[dict[str, Any]]:
        """分页获取对象的所有授权账户."""
        all_accounts: list[dict[str, Any]] = []
        page = 1
        size = 100
        total = 0

        while True:
            accounts, page_info = _get_obj_access_list(
                self.client, obj_id, obj_type, page, size
            )
            all_accounts.extend(accounts)
            total = int(page_info.get("list_total_num", 0) or total)
            if not accounts or len(all_accounts) >= total:
                break
            page += 1
            if page > 50:
                break

        return all_accounts

    @staticmethod
    def _account_display_name(account: dict[str, Any]) -> str:
        """根据账户类型获取显示名称."""
        account_type = account.get("account_type", "user")
        if account_type == "user":
            return account.get("user_name", "") or account.get("account", "")
        if account_type == "class":
            return account.get("class_name", "") or account.get("account", "")
        if account_type == "department":
            return account.get("department_name", "") or account.get("account", "")
        if account_type == "group":
            return account.get("group_name", "") or account.get("account", "")
        return account.get("account", "")

    # -----------------------------------------------------------------------
    # 场景模板：管理员账号列表导出
    # -----------------------------------------------------------------------
    def export_admin_accounts(
        self,
        output_path: str,
        *,
        keywords: str | None = None,
        group_ids: str | None = None,
        group_operator: str = "intersection",
        role_type: int | None = None,
        account_status: int | None = None,
        is_manager: int = 0,
    ) -> dict[str, Any]:
        """导出企业账号列表.

        Args:
            output_path: 输出文件路径，根据扩展名自动选择 xlsx/csv。
            keywords: 搜索关键词（姓名、邮箱、手机号、用户名）。
            group_ids: 分组ID列表，多个用逗号分隔。
            group_operator: 多分组关系，intersection 或 union。
            role_type: 角色筛选：1=学员, 2=讲师, 3=学习负责人, 4=系统管理员, 5=子管理员。
            account_status: 状态筛选：0=待加入, 1=已启用, 2=已禁用, 3=定时禁用。
            is_manager: 0=返回全部账号，1=仅返回管理视角账号。

        Returns:
            包含 file_path、total_records 的字典。
        """
        records = self._fetch_all_admin_accounts(
            keywords=keywords,
            group_ids=group_ids,
            group_operator=group_operator,
            role_type=role_type,
            account_status=account_status,
            is_manager=is_manager,
        )

        ext = os.path.splitext(output_path)[1].lower()
        if ext == ".csv":
            self.to_csv(records, output_path)
        else:
            self.to_excel(records, output_path, sheet_name="企业账号列表")

        return {
            "file_path": output_path,
            "total_records": len(records),
        }

    def _fetch_all_admin_accounts(
        self,
        *,
        keywords: str | None = None,
        group_ids: str | None = None,
        group_operator: str = "intersection",
        role_type: int | None = None,
        account_status: int | None = None,
        is_manager: int = 0,
    ) -> list[dict[str, Any]]:
        """分页获取企业账号列表."""
        all_accounts: list[dict[str, Any]] = []
        page = 1
        size = 500
        total = 0

        while True:
            params: dict[str, str] = {
                "is_manager": str(is_manager),
                "page": str(page),
                "size": str(size),
                "group_operator": group_operator,
            }
            if keywords:
                params["keywords"] = keywords
            if group_ids:
                params["group_ids"] = group_ids
            if role_type is not None:
                params["role_type"] = str(role_type)
            if account_status is not None:
                params["account_status"] = str(account_status)

            resp = self.client.get(
                self.client.desktop_url("/ajax/enterprise/getUserList"),
                params=params,
            )
            if resp.get("status") not in (True, "true") and resp.get("error_code") != 0:
                raise RuntimeError(resp.get("error", "获取账号列表失败"))

            user_list = resp.get("data", {}).get("list", [])
            total = int(
                resp.get("data", {}).get("page_info", {}).get("list_total_num", 0) or total
            )
            all_accounts.extend(user_list)

            if not user_list or len(all_accounts) >= total:
                break
            page += 1
            if page > 50:
                break

        return all_accounts

    # -----------------------------------------------------------------------
    # 场景模板：学习记录导出
    # -----------------------------------------------------------------------
    def export_learning_records(
        self,
        output_path: str,
        *,
        start_day: str | None = None,
        end_day: str | None = None,
        uids: list[str] | None = None,
        course_title: str | None = None,
        department_ids: str | None = None,
        group_ids: str | None = None,
        class_ids: list[str] | None = None,
    ) -> dict[str, Any]:
        """导出企业账号的课程学习明细.

        Args:
            output_path: 输出文件路径，根据扩展名自动选择 xlsx/csv。
            start_day: 最后学习时间起始日期 YYYY-MM-DD。
            end_day: 最后学习时间结束日期 YYYY-MM-DD。
            uids: 学员 UMU ID 数组。
            course_title: 课程名称模糊搜索关键词。
            department_ids: 部门 ID 逗号分隔字符串。
            group_ids: 企业分组 ID 逗号分隔字符串。
            class_ids: 班级 ID 列表。

        Returns:
            包含 file_path、total_records 的字典。
        """
        records = self._fetch_all_learning_records(
            start_day=start_day,
            end_day=end_day,
            uids=uids,
            course_title=course_title,
            department_ids=department_ids,
            group_ids=group_ids,
            class_ids=class_ids,
        )

        ext = os.path.splitext(output_path)[1].lower()
        if ext == ".csv":
            self.to_csv(records, output_path)
        else:
            self.to_excel(records, output_path, sheet_name="学习记录")

        return {
            "file_path": output_path,
            "total_records": len(records),
        }

    def _fetch_all_learning_records(
        self,
        *,
        start_day: str | None = None,
        end_day: str | None = None,
        uids: list[str] | None = None,
        course_title: str | None = None,
        department_ids: str | None = None,
        group_ids: str | None = None,
        class_ids: list[str] | None = None,
    ) -> list[dict[str, Any]]:
        """分页获取学习记录."""
        condition: dict[str, Any] = {}
        if start_day:
            condition["start_date"] = start_day
        if end_day:
            condition["end_date"] = end_day
        if uids:
            condition["uids"] = uids
        if course_title:
            condition["group_title"] = course_title
        if department_ids:
            condition["department_ids"] = [d.strip() for d in department_ids.split(",") if d.strip()]
        if group_ids:
            condition["enterprise_group_ids"] = [g.strip() for g in group_ids.split(",") if g.strip()]
        if class_ids:
            condition["class_ids"] = class_ids

        all_records: list[dict[str, Any]] = []
        page = 1
        size = 100
        total = 0

        while True:
            params: dict[str, str] = {
                "t": str(int(time.time() * 1000)),
                "page": str(page),
                "size": str(size),
                "search_condition": json.dumps(condition, ensure_ascii=False),
            }
            if start_day:
                params["start_day"] = start_day
            if end_day:
                params["end_day"] = end_day
            if department_ids:
                params["department_ids"] = department_ids
            if group_ids:
                params["enterprise_group_ids"] = group_ids
            if class_ids:
                params["class_ids"] = ",".join(class_ids)

            resp = self.client.get(
                self.client.desktop_url("/uapi/v1/dashboard/learning-group-list"),
                params=params,
            )
            if resp.get("error_code") != 0:
                raise RuntimeError(resp.get("error_message", "获取学习记录失败"))

            record_list = resp.get("data", {}).get("list", [])
            total = int(
                resp.get("data", {}).get("page_info", {}).get("list_total_num", 0) or total
            )
            all_records.extend(record_list)

            if not record_list or len(all_records) >= total:
                break
            page += 1
            if page > 50:
                break

        return all_records

    # -----------------------------------------------------------------------
    # 场景模板：学习项目权限导出
    # -----------------------------------------------------------------------
    def export_program_permissions(
        self,
        output_path: str,
        *,
        scope: str = "owned",
        keywords: str | None = None,
        include_access_list: bool = True,
    ) -> dict[str, Any]:
        """导出当前讲师的学习项目访问权限明细.

        Args:
            output_path: 输出文件路径，根据扩展名自动选择 xlsx/csv。
            scope: 列表视角：owned=我拥有的, cooperated=协同给我的, enrolled=我报名的。
            keywords: 按标题/访问码模糊搜索。
            include_access_list: 是否展开指定账户的授权列表。

        Returns:
            包含 file_path、total_programs、total_records 的字典。
        """
        programs = self._fetch_all_programs(scope=scope, keywords=keywords or "")
        records: list[dict[str, Any]] = []

        for program in programs:
            program_id = program.get("program_id", "")
            title = program.get("program_title", "")
            access_code = program.get("access_code", "")

            try:
                access_permission, _, _ = _get_obj_access_permission(
                    self.client, program_id, "program"
                )
                permission_text = _permission_text(access_permission)
            except Exception as e:
                records.append(
                    {
                        "program_id": program_id,
                        "title": title,
                        "access_code": access_code,
                        "access_permission": "",
                        "permission_text": f"获取失败: {e}",
                        "authorized_account": "",
                        "authorized_account_type": "",
                        "authorized_account_id": "",
                        "authorized_account_name": "",
                    }
                )
                continue

            if access_permission == 3 and include_access_list:
                accounts = self._fetch_all_access_list(program_id, "program")
                if accounts:
                    for account in accounts:
                        records.append(
                            {
                                "program_id": program_id,
                                "title": title,
                                "access_code": access_code,
                                "access_permission": access_permission,
                                "permission_text": permission_text,
                                "authorized_account": account.get("account", ""),
                                "authorized_account_type": account.get("account_type", ""),
                                "authorized_account_id": account.get("id", ""),
                                "authorized_account_name": self._account_display_name(account),
                            }
                        )
                else:
                    records.append(
                        {
                            "program_id": program_id,
                            "title": title,
                            "access_code": access_code,
                            "access_permission": access_permission,
                            "permission_text": permission_text,
                            "authorized_account": "",
                            "authorized_account_type": "",
                            "authorized_account_id": "",
                            "authorized_account_name": "",
                        }
                    )
            else:
                records.append(
                    {
                        "program_id": program_id,
                        "title": title,
                        "access_code": access_code,
                        "access_permission": access_permission,
                        "permission_text": permission_text,
                        "authorized_account": "",
                        "authorized_account_type": "",
                        "authorized_account_id": "",
                        "authorized_account_name": "",
                    }
                )

        ext = os.path.splitext(output_path)[1].lower()
        if ext == ".csv":
            self.to_csv(records, output_path)
        else:
            self.to_excel(records, output_path, sheet_name="学习项目访问权限")

        return {
            "file_path": output_path,
            "total_programs": len(programs),
            "total_records": len(records),
        }

    def _fetch_all_programs(
        self,
        *,
        scope: str = "owned",
        keywords: str = "",
    ) -> list[dict[str, Any]]:
        """分页获取学习项目列表."""
        base_params: dict[str, str] = {
            "t": str(int(time.time() * 1000)),
            "page": "1",
            "size": "20",
        }
        if scope == "owned":
            url = "/api/program/getlist"
            base_params["owner"] = "1"
            base_params["type"] = "1"
        elif scope == "cooperated":
            url = "/api/program/getcooperateprogramlist"
        elif scope == "enrolled":
            url = "/api/program/getmyparticipatedprogramlist"
        else:
            raise ValueError(f"不支持的 scope: {scope}")

        if keywords:
            base_params["keywords"] = keywords

        all_items: list[dict[str, Any]] = []
        page = 1
        size = 20
        total = 0

        while True:
            params = {**base_params, "page": str(page), "size": str(size)}
            resp = self.client.get(self.client.desktop_url(url), params=params)
            if resp.get("status") not in (True, "true") and resp.get("error_code") != 0:
                raise RuntimeError(resp.get("error", "获取学习项目列表失败"))

            data = resp.get("data", {})
            page_info = data.get("page_info", {})
            program_list = data.get("list", [])
            total = int(page_info.get("list_total_num", 0) or total)

            for item in program_list:
                creator = item.get("creator", {}) or {}
                all_items.append(
                    {
                        "program_id": str(item.get("program_id", "")),
                        "program_title": item.get("program_title", ""),
                        "access_code": item.get("access_code", ""),
                        "creator_umu_id": str(creator.get("umu_id", "")),
                        "creator_name": creator.get("user_name", item.get("creater_name", "")),
                    }
                )

            if not program_list or len(all_items) >= total:
                break
            page += 1
            if page > 50:
                break

        return all_items
